from django.shortcuts import render
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import  IncomeCategory, ExpenseCategory
from accounts.models import Farm
from .forms import IncomeCategoryForm, ExpenseCategoryForm
from datetime import date, datetime, timedelta
from django.db.models import F
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .serializers import IncomeSerializer

# Other imports and views

@login_required
def income_categories(request):
    farm = request.user.farm
    income_categories = IncomeCategory.objects.filter(farm=farm)

    return render(request, 'farm_finances/income_categories.html', {'farm': farm, 'income_categories': income_categories})

@login_required
def create_income_category(request):
    farm = request.user.farm
    if request.method == 'POST':
        form = IncomeCategoryForm(request.POST)
        if form.is_valid():
            income_category = form.save(commit=False)
            income_category.farm = farm
            income_category.save()
            return redirect('farm_finances:income_categories')
    else:
        form = IncomeCategoryForm()

    return render(request, 'farm_finances/create_income_category.html', {'form': form, 'farm': farm})

@login_required
def create_expense_category(request):
    farm = request.user.farm
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            expense_category = form.save(commit=False)
            expense_category.farm = farm
            expense_category.save()
            return redirect('farm_finances:expense_categories')
    else:
        form = ExpenseCategoryForm()

    return render(request, 'farm_finances/create_expense_category.html', {'form': form, 'farm': farm})

@login_required
def update_income_category(request, income_category_id):
    farm = request.user.farm
    income_category = get_object_or_404(IncomeCategory, pk=income_category_id, farm=farm)
    if request.method == 'POST':
        form = IncomeCategoryForm(request.POST, instance=income_category)
        if form.is_valid():
            form.save()
            return redirect('farm_finances:income_categories')
    else:
        form = IncomeCategoryForm(instance=income_category)
    return render(request, 'farm_finances/update_income_category.html', {'form': form, 'farm': farm, 'income_category': income_category})


@login_required
def update_expense_category(request, expense_category_id):
    farm = request.user.farm
    expense_category = get_object_or_404(ExpenseCategory, pk=expense_category_id, farm=farm)
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST, instance=expense_category)
        if form.is_valid():
            form.save()
            return redirect('farm_finances:expense_categories')
    else:
        form = ExpenseCategoryForm(instance=expense_category)
    return render(request, 'farm_finances/update_expense_category.html', {'form': form, 'farm': farm, 'expense_category': expense_category})

@login_required
def delete_income_category(request, pk):
    income_category = get_object_or_404(IncomeCategory, pk=pk, farm=request.user.farm)
    
    if income_category.name == 'Milk Sale':
        # Prevent deleting Milk Sale category
        return redirect('farm_finances:income_categories')
    
    income_category.delete()
    return redirect('farm_finances:income_categories')



@login_required
def delete_expense_category(request, pk):
    expense_category = get_object_or_404(ExpenseCategory, pk=pk, farm=request.user.farm)
    if expense_category.name == 'Salary':
        return redirect('farm_finances:expense_categories')
    expense_category.delete()
    return redirect('farm_finances:expense_categories')






@login_required
def expense_categories(request):
    farm = request.user.farm
    expense_categories = ExpenseCategory.objects.filter(farm=farm)
    print(expense_categories)

    return render(request, 'farm_finances/expense_categories.html', {'farm': farm, 'expense_categories': expense_categories})


from django.http import HttpResponse, HttpResponseRedirect, JsonResponse



from .models import Expense, ExpenseCategory
from .forms import ExpenseForm, IncomeForm

@login_required
def create_income(request):
    farm = request.user.farm
    if request.method == "POST":
        form = IncomeForm(request.POST, request.FILES, farm=farm)
        if form.is_valid():
            income = form.save(commit=False)
            income.user = request.user
            income.farm = farm
            income.save()
            return redirect('farm_finances:income_list')
    else:
        form = IncomeForm(farm=farm)
        today_str = date.today().strftime('%Y-%m-%d')
    return render(request, 'farm_finances/create_income.html', {'form': form, 'today': today_str})


@login_required
def update_income(request, income_id):
    income = get_object_or_404(Income, id=income_id, user=request.user)
    if request.method == "POST":
        form = IncomeForm(request.POST, request.FILES, instance=income, farm=request.user.farm)
        if form.is_valid():
            form.save()
            return redirect('farm_finances:income_list')
    else:
        form = IncomeForm(instance=income, farm=request.user.farm)
    today_str = date.today().strftime('%Y-%m-%d')
    return render(request, 'farm_finances/update_income.html', {'form': form, 'income': income, 'today': today_str})




@login_required
def delete_income(request, pk):
    income = get_object_or_404(Income, pk=pk, farm=request.user.farm)
    income.delete()
    return redirect('farm_finances:income_list')


from django.shortcuts import render
from .models import Income



from datetime import timedelta, datetime

@login_required
def income_list(request):
    farm = request.user.farm
    sort_by = request.GET.get('sort_by', 'date')
    sort_order = request.GET.get('sort_order', 'desc')
    if sort_order == 'asc':
        incomes = Income.objects.filter(farm=farm).order_by(sort_by)
    else:
        incomes = Income.objects.filter(farm=farm).order_by(F(sort_by).desc(nulls_last=True))

    time_filter = request.GET.get('time_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if time_filter == 'last_7_days':
        start_date = datetime.today() - timedelta(days=7)
        end_date = datetime.today()
    elif time_filter == 'one_month':
        start_date = datetime.today() - timedelta(days=30)
        end_date = datetime.today()
    elif time_filter == 'three_months':
        start_date = datetime.today() - timedelta(days=90)
        end_date = datetime.today()
    elif time_filter == 'one_year':
        start_date = datetime.today() - timedelta(days=365)
        end_date = datetime.today()

    if start_date and end_date:
        incomes = incomes.filter(date__range=(start_date, end_date))

    return render(request, 'farm_finances/income_list.html', {'incomes': incomes, 'farm': farm, 'sort_by': sort_by, 'sort_order': sort_order})


@login_required
def create_expense(request):
    farm = request.user.farm
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES, farm=farm)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.farm = farm
            expense.save()
            return redirect('farm_finances:expense_list')
    else:
        form = ExpenseForm(farm=farm)
    today_str = date.today().strftime('%Y-%m-%d')
    return render(request, 'farm_finances/create_expense.html', {'form': form, 'today': today_str})



from django.shortcuts import render
from .models import Expense

@login_required
def update_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    farm = expense.farm  # get the farm from the expense instance
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES, instance=expense, farm=farm)
        if form.is_valid():
            form.save()
            return redirect('farm_finances:expense_list')
    else:
        form = ExpenseForm(instance=expense, farm=farm)
    today_str = date.today().strftime('%Y-%m-%d')
    return render(request, 'farm_finances/update_expense.html', {'form': form, 'expense': expense, 'today': today_str})

@login_required
def delete_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk, farm=request.user.farm)
    expense.delete()
    return redirect('farm_finances:expense_list')


from datetime import timedelta, datetime
from django.utils import timezone


@login_required
def expense_list(request):
    farm = request.user.farm
    sort_by = request.GET.get('sort_by', 'date')
    sort_order = request.GET.get('sort_order', 'desc')

    if sort_order == 'asc':
        expenses = Expense.objects.filter(farm=farm).order_by(sort_by)
    else:
        expenses = Expense.objects.filter(farm=farm).order_by(F(sort_by).desc(nulls_last=True))

    time_filter = request.GET.get('time_filter', 'one_month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if time_filter == 'last_7_days':
        start_date = datetime.today() - timedelta(days=7)
        end_date = datetime.today()
    elif time_filter == 'one_month':
        start_date = datetime.today() - timedelta(days=30)
        end_date = datetime.today()
    elif time_filter == 'three_months':
        start_date = datetime.today() - timedelta(days=90)
        end_date = datetime.today()
    elif time_filter == 'one_year':
        start_date = datetime.today() - timedelta(days=365)
        end_date = datetime.today()

    if start_date and end_date:
        expenses = expenses.filter(date__range=(start_date, end_date))

    return render(request, 'farm_finances/expense_list.html', {'expenses': expenses, 'farm': farm, 'sort_by': sort_by, 'sort_order': sort_order})


import openpyxl
from .models import Income, Expense
from django.db import models


from django.db.models.fields.files import ImageFieldFile


def export_to_excel(model):
    # create a new excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # get the model's fields and append them as the header row
    fields = model._meta.fields
    header_row = [field.name for field in fields if not isinstance(field, models.ImageField)]
    ws.append(header_row)

    # get the model's objects and append each object as a row
    for obj in model.objects.all():
        row = []
        for field in fields:
            if isinstance(field, models.ImageField):
                continue
            value = getattr(obj, field.name)
            if isinstance(value, models.Model):
                value = str(value)
            elif isinstance(value, ImageFieldFile):
                if value:
                    value = str(value.path)
                else:
                    value = None
            row.append(value)
        ws.append(row)

    return wb




from django.core.exceptions import ObjectDoesNotExist

def import_from_excel(model, file, user):  # add the user parameter
    # open the workbook
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # get the header row and model fields
    header_row = [cell.value for cell in ws[1]]
    fields = {field.name: field for field in model._meta.fields}

    # iterate over the rows in the worksheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        obj_dict = {}
        for key, value in zip(header_row, row):
            field = fields.get(key)
            if field is None:
                continue

            # if the field is a ForeignKey and not the 'user' field, look up the related object
            if isinstance(field, models.ForeignKey) and field.name != 'user':
                # check which field to use for lookup based on field name
                lookup_field = 'name' if field.name in ['farm', 'category'] else 'id'  # revise this as needed
                try:
                    value = field.related_model.objects.get(**{lookup_field: value})
                except ObjectDoesNotExist:
                    value = None
            obj_dict[key] = value

        # assign the logged-in user to the 'user' field
        obj_dict['user'] = user

        # create the object
        model.objects.create(**obj_dict)




@login_required
def export_income(request):
    wb = export_to_excel(Income)
    # Save the workbook to a HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Incomes.xlsx'
    wb.save(response)
    return response

@login_required
def export_expense(request):
    wb = export_to_excel(Expense)
    # Save the workbook to a HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Expenses.xlsx'
    wb.save(response)
    return response


@login_required
def import_income(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            import_from_excel(Income, request.FILES['file'], request.user)  # pass the user here
            return HttpResponseRedirect('/success/url/')
    else:
        form = UploadFileForm()
    return render(request, 'farm_finances/upload.html', {'form': form})

@login_required
def import_expense(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            import_from_excel(Expense, request.FILES['file'], request.user)
            return HttpResponseRedirect('/success/url/')
    else:
        form = UploadFileForm()
    return render(request, 'farm_finances/upload.html', {'form': form})

from django import forms

class UploadFileForm(forms.Form):
    file = forms.FileField()

# Mobile API

from django.db.models import Sum, F
from django.utils import timezone
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import viewsets
from .serializers import IncomeSerializer
from .models import Income, IncomeCategory, Farm

class IncomeViewSet(viewsets.ModelViewSet):
    queryset = Income.objects.all()
    serializer_class = IncomeSerializer

    @action(detail=False, methods=['get'])
    def income_list(self, request):
        user = request.user
        farm = user.farm
        sort_by = request.GET.get('sort_by', 'date')
        sort_order = request.GET.get('sort_order', 'desc')
        if sort_order == 'asc':
            incomes = Income.objects.filter(farm=farm).order_by(sort_by)
        else:
            incomes = Income.objects.filter(farm=farm).order_by(F(sort_by).desc(nulls_last=True))

        time_filter = request.GET.get('time_filter')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if time_filter == 'last_7_days':
            start_date = timezone.now() - timezone.timedelta(days=7)
            end_date = timezone.now()
        elif time_filter == 'one_month':
            start_date = timezone.now() - timezone.timedelta(days=30)
            end_date = timezone.now()
        elif time_filter == 'three_months':
            start_date = timezone.now() - timezone.timedelta(days=90)
            end_date = timezone.now()
        elif time_filter == 'one_year':
            start_date = timezone.now() - timezone.timedelta(days=365)
            end_date = timezone.now()

        if start_date and end_date:
            incomes = incomes.filter(date__range=(start_date, end_date))

        income_categories = IncomeCategory.objects.filter(farm=farm)

        summary = []
        for category in income_categories:
            total_amount = incomes.filter(category=category).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
            summary.append({
                "category": category.name,
                "total_amount": total_amount,
            })

        serializer = IncomeSerializer(incomes, many=True)
        return Response({
            'income_list': serializer.data,
            'summary': summary
        })
